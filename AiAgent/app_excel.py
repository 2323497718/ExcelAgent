"""
Streamlit UI for Excel Agent.
Left sidebar: Excel file browser + data editor
Main area: Agent chat (fragment handles async tool execution)
"""

import sys
import os

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

import streamlit as st
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from langchain_core.messages import HumanMessage, AIMessage

from core.tools.excel_tools import (
    excel_read_file,
    excel_read_sheet_data,
    excel_create_file,
    excel_write_cells,
    excel_write_row,
    excel_append_row,
    excel_add_sheet,
    excel_delete_rows,
)

from core.helper.llm_util import init_llm
from core.excel_agent_builder import ExcelAgent


# ══════════════════════════════════════════════════════════════
#  Page Config
# ══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Excel 智能体",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ══════════════════════════════════════════════════════════════
#  CSS — Dark theme
# ══════════════════════════════════════════════════════════════
st.markdown("""
<style>
    :root {
        --bg-primary:    #0d1117;
        --bg-secondary:  #161b22;
        --bg-card:       #1c2128;
        --bg-hover:      #21262d;
        --border:        #30363d;
        --text-primary:  #e6edf3;
        --text-muted:    #8b949e;
        --accent-blue:   #58a6ff;
        --accent-green:  #3fb950;
        --accent-red:    #f85149;
    }

    .stApp { background: var(--bg-primary); }

    [data-testid="stSidebar"] {
        background: var(--bg-secondary) !important;
        border-right: 1px solid var(--border);
        min-width: 340px !important;
    }

    /* Metrics */
    [data-testid="stMetric"] {
        background: var(--bg-card);
        border: 1px solid var(--border);
        border-radius: 8px;
        padding: 8px 12px;
    }
    [data-testid="stMetricLabel"] { color: var(--text-muted) !important; font-size: 0.68rem !important; }
    [data-testid="stMetricValue"]  { color: var(--accent-blue) !important; font-size: 1.2rem !important; }

    /* Buttons */
    .stButton > button {
        border: 1px solid var(--border) !important;
        background: var(--bg-hover) !important;
        color: var(--text-primary) !important;
        border-radius: 6px !important;
        font-size: 0.8rem !important;
        transition: all 0.15s !important;
        width: 100% !important;
    }
    .stButton > button:hover {
        background: #2d333b !important;
        border-color: var(--accent-blue) !important;
        color: var(--accent-blue) !important;
    }
    /* Primary button */
    [data-testid="stBaseButton-primary"] {
        background: var(--accent-blue) !important;
        color: #fff !important;
        border: none !important;
    }

    /* Text inputs */
    [data-testid="stTextInput"] label,
    [data-testid="stNumberInput"] label { color: var(--text-muted) !important; font-size: 0.72rem !important; }

    /* File uploader */
    [data-testid="stFileUploaderDropzone"] {
        border: 2px dashed #58a6ff !important;
        border-radius: 10px !important;
        background: rgba(88,166,255,0.06) !important;
    }
    [data-testid="stFileUploader"] label { color: var(--accent-blue) !important; font-size: 0.85rem !important; font-weight: 600 !important; }

    /* Divider */
    hr { border-color: var(--border) !important; }

    /* Chat bubbles */
    [data-testid="stChatMessageUser"] {
        background: #1a3a5c !important;
        border: 1px solid #2a5f8f !important;
        border-radius: 14px 14px 4px 14px !important;
        padding: 10px 14px !important;
        max-width: 82% !important;
        margin-left: auto !important;
    }
    [data-testid="stChatMessage"] { background: transparent !important; padding: 3px 0 !important; }

    /* Tool result card */
    .tool-result {
        background: #0d1117;
        border: 1px solid #30363d;
        border-left: 3px solid #58a6ff;
        border-radius: 6px;
        padding: 8px 12px;
        margin: 4px 0;
        font-size: 0.78rem;
        color: #8b949e;
        white-space: pre-wrap;
        word-break: break-all;
        max-height: 220px;
        overflow-y: auto;
    }

    /* Status indicator */
    .status-ready { display: inline-flex; align-items: center; gap: 5px; font-size: 0.78rem; color: var(--accent-green); }
    .status-dot { width: 8px; height: 8px; border-radius: 50%; background: var(--accent-green); box-shadow: 0 0 6px var(--accent-green); animation: pulse 2s infinite; }
    @keyframes pulse { 0%,100%{opacity:1;} 50%{opacity:0.4;} }

    /* File badge */
    .file-badge { display: inline-flex; align-items: center; gap: 4px; background: rgba(88,166,255,0.12); border: 1px solid rgba(88,166,255,0.3); border-radius: 20px; padding: 4px 12px; font-size: 0.78rem; color: #79b8ff; }

    /* Section label */
    .sidebar-section { color: var(--text-muted); font-size: 0.65rem; text-transform: uppercase; letter-spacing: 0.1em; font-weight: 600; margin: 10px 0 3px 0; }

    /* Thinking spinner */
    .thinking-dot { animation: blink 1s infinite; display: inline-block; }
    @keyframes blink { 0%,100%{opacity:1;} 50%{opacity:0.3;} }

    /* Hide branding */
    #MainMenu, footer, header { visibility: hidden; }

    /* Scrollbar */
    ::-webkit-scrollbar { width: 5px; height: 5px; }
    ::-webkit-scrollbar-track { background: var(--bg-primary); }
    ::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }
    ::-webkit-scrollbar-thumb:hover { background: #484f58; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
#  Helpers
# ══════════════════════════════════════════════════════════════
def load_excel_sheet(path: str, sheet: str = None):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        ws = wb.active
    sheets = wb.sheetnames
    active = ws.title
    headers, rows = [], []
    if ws:
        max_col = ws.max_column or 0
        headers = [str(ws.cell(1, c).value or "") for c in range(1, max_col + 1)]
        for r in range(2, ws.max_row + 1):
            row = [str(ws.cell(r, c).value or "") for c in range(1, max_col + 1)]
            rows.append(row)
    wb.close()
    return sheets, active, headers, rows


def detect_changes(old_df: pd.DataFrame, new_df: pd.DataFrame) -> list:
    changes = []
    for col_idx in range(len(old_df.columns)):
        for row_idx in range(len(old_df)):
            old_v = str(old_df.iloc[row_idx, col_idx] or "")
            new_v = str(new_df.iloc[row_idx, col_idx] or "")
            if old_v != new_v:
                addr = f"{get_column_letter(col_idx + 1)}{row_idx + 2}"
                changes.append((addr, new_v))
    return changes


def execute_tool(name: str, args: dict) -> str:
    tool_map = {
        "excel_read_file":       excel_read_file,
        "excel_read_sheet_data": excel_read_sheet_data,
        "excel_create_file":     excel_create_file,
        "excel_write_cells":     excel_write_cells,
        "excel_write_row":       excel_write_row,
        "excel_append_row":      excel_append_row,
        "excel_add_sheet":       excel_add_sheet,
        "excel_delete_rows":     excel_delete_rows,
    }
    fn = tool_map.get(name)
    if not fn:
        return f"Unknown tool: {name}"
    try:
        return fn.invoke(args)
    except Exception as e:
        return f"Error: {str(e)}"


# ══════════════════════════════════════════════════════════════
#  Session State defaults
# ══════════════════════════════════════════════════════════════
for k, v in {
    "file_path":    "",
    "sheets":       [],
    "active_sheet": "",
    "headers":      [],
    "rows":         [],
    "messages":     [],       # {"role": "user"|"assistant"|"tool", "content"|"tool"|"result": ...}
    "agent_phase":  "idle",   # idle | running
    "agent_ready":  False,
    "uploaded_name": None,
}.items():
    st.session_state.setdefault(k, v)


# ══════════════════════════════════════════════════════════════
#  LLM / Agent (cached)
# ══════════════════════════════════════════════════════════════
@st.cache_resource
def get_agent():
    llm = init_llm(api_key="sk-5588ad0c13e44635bbdddac949f1e874")
    return ExcelAgent(llm=llm, verbose=False)

agent = get_agent()
st.session_state.agent_ready = True


# ══════════════════════════════════════════════════════════════
#  SIDEBAR — File Panel
# ══════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### 📁 Excel 文件")

    # Upload
    st.markdown('<p class="sidebar-section">上传文件</p>', unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "拖拽或点击上传 Excel 文件",
        type=["xlsx", "xlsm", "xls"],
        key="file_uploader_widget",
    )
    if uploaded is not None and st.session_state.get("uploaded_name") != uploaded.name:
        save_dir = os.path.join("output", "uploaded")
        os.makedirs(save_dir, exist_ok=True)
        save_path = os.path.join(save_dir, uploaded.name)
        with open(save_path, "wb") as f:
            f.write(uploaded.getbuffer())
        st.session_state.file_path = save_path
        st.session_state.active_sheet = ""
        st.session_state.uploaded_name = uploaded.name

    # Path input
    st.markdown('<p class="sidebar-section">或输入路径</p>', unsafe_allow_html=True)
    raw_path = st.text_input(
        "文件路径",
        value=st.session_state.file_path,
        placeholder="C:/data/report.xlsx",
        label_visibility="collapsed",
    )
    if raw_path and raw_path != st.session_state.file_path:
        st.session_state.file_path = raw_path
        st.session_state.active_sheet = ""

    # File viewer
    fp = st.session_state.file_path
    if fp and os.path.exists(fp):
        try:
            sheets, active, headers, rows = load_excel_sheet(fp, st.session_state.active_sheet or None)
            st.session_state.sheets = sheets

            st.markdown(f"<span class='file-badge'>📄 {os.path.basename(fp)}</span>", unsafe_allow_html=True)

            sel_idx = 0
            if st.session_state.active_sheet and st.session_state.active_sheet in sheets:
                sel_idx = sheets.index(st.session_state.active_sheet)
            sel = st.selectbox("工作表", sheets, index=sel_idx)
            if sel != st.session_state.active_sheet:
                st.session_state.active_sheet = sel
                _, _, headers, rows = load_excel_sheet(fp, sel)

            cr, cc, cs = st.columns(3)
            cr.metric("行", len(rows) + 1)
            cc.metric("列", len(headers))
            cs.metric("工作表", len(sheets))

            st.markdown("---")

            if headers:
                st.markdown('<p class="sidebar-section">数据预览</p>', unsafe_allow_html=True)
                df = pd.DataFrame(rows, columns=headers)
                df.index = pd.Index(range(2, len(rows) + 2), name="行号")

                edited_df = st.data_editor(
                    df,
                    use_container_width=True,
                    num_rows="dynamic",
                    key="data_editor",
                    column_config={h: st.column_config.TextColumn(h) for h in headers},
                    hide_index=False,
                )

                changes = detect_changes(df, edited_df)
                save_label = f"💾 保存修改 ({len(changes)})" if changes else "💾 保存修改"
                if st.button(save_label, use_container_width=True):
                    if changes:
                        for addr, val in changes:
                            excel_write_cells.invoke({
                                "file_path": fp, "cell": addr, "value": val, "sheet_name": sel
                            })
                        st.success(f"✅ 已保存 {len(changes)} 处修改")
                    else:
                        st.info("没有检测到修改")

                st.markdown("---")

                # Quick actions
                st.markdown('<p class="sidebar-section">快捷操作</p>', unsafe_allow_html=True)
                qc1, qc2 = st.columns(2)
                with qc1:
                    if st.button("➕ 追加行", use_container_width=True):
                        excel_append_row.invoke({
                            "file_path": fp, "values": [""] * len(headers), "sheet_name": sel
                        })
                        st.success("已追加空行")
                with qc2:
                    new_s = st.text_input("新工作表名", key="new_sheet_in", label_visibility="collapsed", placeholder="新工作表名")
                    if st.button("📑 新建工作表", use_container_width=True) and new_s:
                        excel_add_sheet.invoke({"file_path": fp, "sheet_name": new_s, "headers": headers})
                        st.success(f"已创建: {new_s}")
                        st.session_state.active_sheet = ""

                del_r = st.number_input("删除行号", min_value=2, value=2, key="del_row_in", label_visibility="collapsed")
                if st.button(f"🗑 删除第 {del_r} 行", use_container_width=True):
                    excel_delete_rows.invoke({"file_path": fp, "row_number": int(del_r), "sheet_name": sel})
                    st.warning(f"已删除第 {del_r} 行")

                st.markdown("---")
                st.caption("💡 对话时智能体会自动读取此文件内容")

        except Exception as e:
            st.error(f"读取失败: {e}")
    else:
        if fp:
            st.error("文件不存在")
        st.info("👆 上传文件或输入路径")


# ══════════════════════════════════════════════════════════════
#  MAIN — Agent Chat
# ══════════════════════════════════════════════════════════════

# Title bar
title_col, status_col = st.columns([1, 4])
with title_col:
    st.markdown("## 🤖 Excel 智能体")
with status_col:
    st.markdown(
        '<span class="status-ready"><span class="status-dot"></span>就绪</span>',
        unsafe_allow_html=True
    )

# File badge
fp = st.session_state.file_path
if fp and os.path.exists(fp):
    st.markdown(
        f"<span class='file-badge'>📎 {os.path.basename(fp)}</span>",
        unsafe_allow_html=True
    )

st.markdown("---")


# ── Agent fragment ───────────────────────────────────────────
#
#  Architecture:
#  - The fragment owns the chat history display and the chat input.
#    (elements inside a fragment don't get cleared between fragment reruns,
#     they accumulate — but since we replace messages[] on each rerun from
#     scratch, the display is always correct.)
#  - When the user submits, we set agent_phase="running" and trigger
#    a fragment rerun. The fragment does ONE step of agent work per rerun,
#    then calls st.rerun(scope="fragment") to continue.
#  - Tool results appear immediately because the fragment reruns, redraws
#    the chat container with the new message, and repeats until done.
#
@st.fragment
def agent_fragment():
    phase = st.session_state.agent_phase

    # ── Render current chat history ──────────────────────────
    for msg in st.session_state.messages:
        if msg["role"] == "user":
            with st.chat_message("user", avatar="👤"):
                st.markdown(msg["content"])
        elif msg["role"] == "assistant":
            with st.chat_message("assistant", avatar="🤖"):
                st.markdown(msg["content"])
        elif msg["role"] == "tool":
            with st.chat_message("assistant", avatar="🤖"):
                st.markdown(f"**🔧 `{msg['tool']}`**")
                st.markdown(
                    f'<div class="tool-result">{msg["result"]}</div>',
                    unsafe_allow_html=True
                )

    # ── Thinking indicator ───────────────────────────────────
    if phase == "running":
        with st.chat_message("assistant", avatar="🤖"):
            col1, col2 = st.columns([6, 1])
            with col1:
                with st.spinner("🤔 智能体思考中…"):
                    pass  # spinner text
            with col2:
                st.markdown('<span class="thinking-dot">●</span>', unsafe_allow_html=True)

    # ── Chat input (always last inside fragment) ─────────────
    if prompt := st.chat_input("向智能体描述你的需求…"):
        st.session_state.messages.append({"role": "user", "content": prompt})
        st.session_state.agent_phase = "running"
        st.rerun(scope="fragment")
        return

    # ── Agent execution loop (step by step) ──────────────────
    if phase != "running":
        return

    fp = st.session_state.file_path

    # Build langchain history
    lang_hist = []
    for m in st.session_state.messages:
        if m["role"] == "user":
            lang_hist.append(HumanMessage(content=m["content"]))
        elif m["role"] in ("assistant", "tool"):
            text = m.get("content") or m.get("result") or ""
            lang_hist.append(AIMessage(content=text))

    # User prompt
    user_prompt = ""
    for m in reversed(st.session_state.messages):
        if m["role"] == "user":
            user_prompt = m["content"]
            break

    # File context
    file_context = ""
    if fp and os.path.exists(fp):
        try:
            info = excel_read_file.invoke({"file_path": fp})
            file_context = f"\n[当前打开的 Excel 文件: {fp}]\n{info}\n"
        except Exception as e:
            file_context = f"\n[文件读取失败: {e}]\n"

    full_input = file_context + user_prompt

    try:
        resp = agent.agent.invoke({"input": full_input, "chat_history": lang_hist})

        resp_text = getattr(resp, "content", str(resp))
        tool_calls = getattr(resp, "tool_calls", None) or []

        if not tool_calls:
            st.session_state.messages.append({"role": "assistant", "content": resp_text})
            st.session_state.agent_phase = "idle"
            st.rerun(scope="fragment")
            return

        # Execute tools
        for tc in tool_calls:
            tn = tc.get("name") or ""
            ta = tc.get("args", {}) or {}
            result = execute_tool(tn, ta)
            st.session_state.messages.append({"role": "tool", "tool": tn, "result": result})
            lang_hist.append(resp)
            lang_hist.append(HumanMessage(content=f"Tool {tn} returned:\n{result}"))

        # Continue to next step
        st.rerun(scope="fragment")

    except Exception as e:
        st.session_state.messages.append({
            "role": "assistant",
            "content": f"❌ 执行出错: {str(e)}"
        })
        st.session_state.agent_phase = "idle"
        st.rerun(scope="fragment")


agent_fragment()


# ── Bottom bar ───────────────────────────────────────────────
st.markdown("---")
bl, br = st.columns([1, 4])
with bl:
    if st.button("🗑 清空对话"):
        st.session_state.messages = []
        st.session_state.agent_phase = "idle"
        st.rerun()
with br:
    st.caption("提示：告诉智能体你想做什么，例如「看看这个表」「把所有空白填上」「加一列备注」")
