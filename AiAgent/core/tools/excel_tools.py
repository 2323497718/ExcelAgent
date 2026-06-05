"""
Excel file reading and manipulation tools using openpyxl.
"""

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')
from typing import Optional, List, Dict, Any
from langchain_core.tools import tool


def _load_workbook(file_path: str):
    """Load an Excel workbook, create if not exists."""
    import openpyxl
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        return wb, True
    wb = openpyxl.load_workbook(file_path, data_only=True)
    return wb, False


def _serialize_value(value: Any) -> str:
    """Serialize a cell value to string for LLM consumption."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)
    return str(value)


def _get_sheet_summary(sheet) -> Dict[str, Any]:
    """Get a summary of a sheet: dimensions, headers, sample rows."""
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0

    if max_row == 0 or max_col == 0:
        return {"row_count": 0, "col_count": 0, "headers": [], "sample_rows": []}

    headers = []
    if max_row >= 1:
        headers = [_serialize_value(sheet.cell(1, c).value) for c in range(1, max_col + 1)]

    sample_rows = []
    sample_count = min(3, max_row - 1)
    for r in range(2, 2 + sample_count):
        row = {}
        for c in range(1, max_col + 1):
            row[headers[c - 1]] = _serialize_value(sheet.cell(r, c).value)
        sample_rows.append(row)

    return {
        "row_count": max_row,
        "col_count": max_col,
        "headers": headers,
        "sample_rows": sample_rows,
    }


@tool("excel_read_file")
def excel_read_file(file_path: str) -> str:
    """
    Read an Excel file and return a structured summary of all sheets.

    Use this as the FIRST step when a user wants to understand an Excel file.
    This returns:
    - Sheet names and their row/col counts
    - Column headers for each sheet
    - Sample rows (up to 3 rows of data per sheet)

    Args:
        file_path: Path to the Excel file (.xlsx, .xlsm, .xls)

    Returns:
        A structured summary of all sheets in the file
    """
    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}"

    try:
        wb, _ = _load_workbook(file_path)
    except Exception as e:
        return f"Error: Could not open file: {str(e)}"

    sheets = {}
    for name in wb.sheetnames:
        sheet = wb[name]
        sheets[name] = _get_sheet_summary(sheet)
    wb.close()

    if not sheets:
        return f"The file '{file_path}' has no sheets."

    lines = [f"# Excel File: {file_path}\n"]
    lines.append(f"**Total sheets**: {len(sheets)}\n")
    lines.append(f"**Sheet names**: {', '.join(wb.sheetnames if False else list(sheets.keys()))}\n")

    for name, info in sheets.items():
        lines.append(f"\n## Sheet: {name}")
        lines.append(f"- Rows: {info['row_count']}, Columns: {info['col_count']}")
        if info['headers']:
            lines.append(f"- **Headers**: `{' | '.join(info['headers'])}`")
        if info['sample_rows']:
            lines.append("\n**Sample data (first 3 data rows, with Excel row numbers):**\n")
            lines.append("| Excel行号 | " + " | ".join(info['headers']) + " |")
            lines.append("| " + " | ".join(["---"] * (len(info['headers']) + 1)) + " |")
            sample_start = 2
            for idx, row in enumerate(info['sample_rows']):
                excel_row = sample_start + idx
                lines.append("| " + str(excel_row) + " | " + " | ".join(str(row.get(h, "")) for h in info['headers']) + " |")
        else:
            lines.append("(no data rows)")

    return "\n".join(lines)


@tool("excel_read_sheet_data")
def excel_read_sheet_data(
    file_path: str,
    sheet_name: Optional[str] = None,
    max_rows: int = 100
) -> str:
    """
    Read detailed data from a specific sheet.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet (defaults to first sheet)
        max_rows: Maximum number of data rows to return (default: 100)

    Returns:
        Full content of the sheet as a table
    """
    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}"

    try:
        wb, _ = _load_workbook(file_path)
    except Exception as e:
        return f"Error: Could not open file: {str(e)}"

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    max_row = min(sheet.max_row or 0, max_rows + 2)
    max_col = sheet.max_column or 0

    if max_row == 0:
        return f"Sheet '{sheet.title}' is empty."

    total_rows = sheet.max_row or 0
    headers = [_serialize_value(sheet.cell(1, c).value) for c in range(1, max_col + 1)]

    lines = [f"# Sheet: {sheet.title}"]
    lines.append(f"Showing rows 1-{max_row} of {total_rows} (row 1 is the header row)\n")
    lines.append("| Excel行号 | " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * (len(headers) + 1)) + " |")

    for r in range(2, max_row + 1):
        row_vals = [_serialize_value(sheet.cell(r, c).value) for c in range(1, max_col + 1)]
        lines.append("| " + str(r) + " | " + " | ".join(row_vals) + " |")

    wb.close()
    return "\n".join(lines)


@tool("excel_read_cell")
def excel_read_cell(file_path: str, cell: str, sheet_name: Optional[str] = None) -> str:
    """
    Read a specific cell or range from a sheet.

    Args:
        file_path: Path to the Excel file
        cell: Cell address (e.g., 'A1', 'B2') or range (e.g., 'A1:C5')
        sheet_name: Name of the sheet (defaults to first sheet)

    Returns:
        The value(s) at the specified cell(s)
    """
    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}"

    try:
        wb, _ = _load_workbook(file_path)
    except Exception as e:
        return f"Error: Could not open file: {str(e)}"

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    try:
        result = sheet[cell]
    except Exception as e:
        return f"Error: Invalid cell reference '{cell}': {str(e)}"

    if hasattr(result, 'value'):
        wb.close()
        return str(result.value) if result.value is not None else "(empty)"

    lines = []
    for row in result:
        row_vals = [_serialize_value(c.value) for c in row]
        lines.append(" | ".join(row_vals))

    wb.close()
    return "\n".join(lines)


@tool("excel_create_file")
def excel_create_file(
    file_path: str,
    headers: List[str],
    data: Optional[List[List[str]]] = None,
    sheet_name: str = "Sheet1"
) -> str:
    """
    Create a new Excel file with the specified headers and optional data.

    Args:
        file_path: Path where the Excel file will be created
        headers: List of column header names
        data: Optional list of rows (each row is a list of cell values)
        sheet_name: Name of the first sheet (default: 'Sheet1')

    Returns:
        Confirmation message with file path and structure
    """
    if os.path.exists(file_path):
        return f"Error: File already exists: {file_path}. Use 'excel_write_cells' to modify existing files."

    if not headers:
        return "Error: At least one header is required."

    dir_path = os.path.dirname(os.path.abspath(file_path))
    if dir_path:
        os.makedirs(dir_path, exist_ok=True)

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx, header)

    if data:
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row_idx, col_idx, value)

    wb.save(file_path)
    wb.close()

    row_count = len(data) if data else 0
    return (
        f"Excel file created successfully!\n"
        f"Path: {file_path}\n"
        f"Sheet: {sheet_name}\n"
        f"Columns: {len(headers)} ({', '.join(headers)})\n"
        f"Data rows: {row_count}"
    )


@tool("excel_write_cells")
def excel_write_cells(
    file_path: str,
    cell: str,
    value: str,
    sheet_name: Optional[str] = None
) -> str:
    """
    Write a value to a specific cell. Creates the file/sheet if they don't exist.

    Args:
        file_path: Path to the Excel file
        cell: Cell address (e.g., 'A1', 'B2')
        value: Value to write
        sheet_name: Target sheet name (defaults to first sheet or 'Sheet1')

    Returns:
        Confirmation with cell address and value written
    """
    import openpyxl

    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name or "Sheet1"
    else:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
            else:
                ws = wb[sheet_name]
        else:
            ws = wb.active

    ws[cell] = value
    wb.save(file_path)
    wb.close()

    return f"Written '{value}' to cell {cell} in sheet '{ws.title}' of {file_path}"


@tool("excel_write_row")
def excel_write_row(
    file_path: str,
    row_number: int,
    values: List[str],
    sheet_name: Optional[str] = None,
    start_col: str = "A"
) -> str:
    """
    Write an entire row of data to a sheet.

    Args:
        file_path: Path to the Excel file
        row_number: Row number to write (1-based)
        values: List of values to write in the row
        sheet_name: Target sheet name (defaults to first sheet)
        start_col: Starting column letter (default: 'A')

    Returns:
        Confirmation with row number and values
    """
    import openpyxl
    from openpyxl.utils import column_index_from_string

    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}. Use 'excel_create_file' first."

    wb = openpyxl.load_workbook(file_path)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        ws = wb[sheet_name]
    else:
        ws = wb.active

    start_col_idx = column_index_from_string(start_col)

    for col_offset, val in enumerate(values):
        ws.cell(row_number, start_col_idx + col_offset, val)

    wb.save(file_path)
    wb.close()

    return (
        f"Written {len(values)} value(s) to row {row_number} "
        f"starting at column {start_col} in sheet '{ws.title}' of {file_path}"
    )


@tool("excel_append_row")
def excel_append_row(
    file_path: str,
    values: List[str],
    sheet_name: Optional[str] = None
) -> str:
    """
    Append a new row of data to the end of a sheet (after the last used row).

    Args:
        file_path: Path to the Excel file
        values: List of values to append
        sheet_name: Target sheet name (defaults to first sheet)

    Returns:
        Confirmation with the row number the data was written to
    """
    import openpyxl

    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}. Use 'excel_create_file' first."

    wb = openpyxl.load_workbook(file_path)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found."
        ws = wb[sheet_name]
    else:
        ws = wb.active

    next_row = ws.max_row + 1
    for col_idx, val in enumerate(values, start=1):
        ws.cell(next_row, col_idx, val)

    wb.save(file_path)
    wb.close()

    return f"Appended {len(values)} value(s) to row {next_row} in sheet '{ws.title}' of {file_path}"


@tool("excel_add_sheet")
def excel_add_sheet(
    file_path: str,
    sheet_name: str,
    headers: Optional[List[str]] = None
) -> str:
    """
    Add a new sheet to an existing Excel file.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name for the new sheet
        headers: Optional list of column headers to write in row 1

    Returns:
        Confirmation with sheet name
    """
    import openpyxl

    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}. Use 'excel_create_file' first."

    wb = openpyxl.load_workbook(file_path)

    if sheet_name in wb.sheetnames:
        wb.close()
        return f"Error: Sheet '{sheet_name}' already exists. Available: {wb.sheetnames}"

    ws = wb.create_sheet(sheet_name)

    if headers:
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(1, col_idx, header)

    wb.save(file_path)
    wb.close()

    return f"Created sheet '{sheet_name}' in {file_path}"


@tool("excel_delete_rows")
def excel_delete_rows(
    file_path: str,
    row_number: int,
    count: int = 1,
    sheet_name: Optional[str] = None
) -> str:
    """
    Delete one or more rows from a sheet.

    Args:
        file_path: Path to the Excel file
        row_number: Starting row number to delete (1-based)
        count: Number of rows to delete (default: 1)
        sheet_name: Target sheet name (defaults to first sheet)

    Returns:
        Confirmation with row numbers deleted
    """
    import openpyxl

    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}"

    wb = openpyxl.load_workbook(file_path)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found."
        ws = wb[sheet_name]
    else:
        ws = wb.active

    ws.delete_rows(row_number, count)
    wb.save(file_path)
    wb.close()

    return f"Deleted {count} row(s) starting from row {row_number} in sheet '{ws.title}'"


@tool("excel_find_rows")
def excel_find_rows(
    file_path: str,
    match_column: Optional[str] = None,
    match_value: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> str:
    """
    Search for rows by a column's value and return the EXACT Excel row numbers.
    This is the MOST IMPORTANT tool for delete/update operations — always use this
    first to find the correct row number before writing or deleting.

    Args:
        file_path: Path to the Excel file
        match_column: Column name (header) or column letter (e.g., 'A' or '姓名')
        match_value: The value to search for (partial match supported)
        sheet_name: Target sheet name (defaults to first sheet)

    Returns:
        A table of matching rows with their Excel row numbers (e.g., row 5 = Excel row 5)
    """
    import openpyxl

    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}"

    wb = openpyxl.load_workbook(file_path, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        ws = wb[sheet_name]
    else:
        ws = wb.active

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    if max_row < 2:
        wb.close()
        return "Sheet is empty or has no data rows."

    headers = [_serialize_value(ws.cell(1, c).value) for c in range(1, max_col + 1)]

    # Resolve column index
    col_idx = None
    if match_column:
        from openpyxl.utils import column_index_from_string
        if match_column.isalpha():
            col_idx = column_index_from_string(match_column)
        else:
            for c in range(1, max_col + 1):
                if headers[c - 1].strip() == match_column.strip():
                    col_idx = c
                    break

    if col_idx is None and match_column:
        wb.close()
        return f"Error: Column '{match_column}' not found. Headers: {headers}"

    results = []
    match_lower = (match_value or "").lower()

    for r in range(2, max_row + 1):
        if match_column and match_value:
            cell_val = ws.cell(r, col_idx).value
            cell_str = (_serialize_value(cell_val) or "").lower()
            if match_lower in cell_str:
                row_data = [_serialize_value(ws.cell(r, c).value) for c in range(1, max_col + 1)]
                results.append((r, row_data))
        elif not match_column:
            for c in range(1, max_col + 1):
                cell_str = (_serialize_value(ws.cell(r, c).value) or "").lower()
                if match_lower in cell_str:
                    row_data = [_serialize_value(ws.cell(r, c_val).value) for c_val in range(1, max_col + 1)]
                    results.append((r, row_data))
                    break

    wb.close()

    if not results:
        return f"No rows found matching '{match_value}' in column '{match_column}'."

    lines = [f"Found {len(results)} matching row(s) (Excel row numbers are in the first column):\n"]
    lines.append("| Excel行号 | " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * (len(headers) + 1)) + " |")
    for excel_row, row_data in results:
        lines.append("| " + str(excel_row) + " | " + " | ".join(str(v) for v in row_data) + " |")

    return "\n".join(lines)
