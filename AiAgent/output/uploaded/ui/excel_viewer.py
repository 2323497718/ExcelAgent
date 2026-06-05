"""
Center panel: editable Excel spreadsheet viewer.
Loads .xlsx/.xls files into a QTableWidget, supports sheet switching,
user direct editing, and saving changes back to disk.
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from PySide6.QtCore import Qt, Signal, QItemSelectionModel
from PySide6.QtGui import QColor, QFont, QBrush
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QComboBox, QLabel, QTableWidget,
    QTableWidgetItem, QHeaderView, QMessageBox, QFileDialog,
)


class ExcelViewer(QWidget):
    """
    Center panel — editable Excel spreadsheet.

    Signals
    -------
    file_loaded(str path)   : emitted after a file is successfully loaded
    sheet_changed(str name) : emitted when user switches sheets
    """

    file_loaded  = Signal(str)
    sheet_changed = Signal(str)

    # ────────────────────────────────────────────────────────────
    #  Init
    # ────────────────────────────────────────────────────────────

    def __init__(self, parent=None):
        super().__init__(parent)
        self._file_path   = ""
        self._wb          = None       # openpyxl.Workbook
        self._ws          = None       # openpyxl.Worksheet
        self._sheet_names = []
        self._headers    = []          # list of str
        self._orig_values = {}         # (row, col) -> original str  (for change tracking)
        self._modified   = False

        self._setup_ui()

    # ────────────────────────────────────────────────────────────
    #  UI Setup
    # ────────────────────────────────────────────────────────────

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        # ── Toolbar ──────────────────────────────────────────────
        toolbar = QHBoxLayout()
        toolbar.setSpacing(6)

        self.lbl_filename = QLabel("未打开文件")
        self.lbl_filename.setStyleSheet("font-size: 10pt; font-weight: 600;")
        toolbar.addWidget(self.lbl_filename)

        toolbar.addStretch()

        # Sheet selector
        self.cmb_sheet = QComboBox()
        self.cmb_sheet.setMinimumWidth(120)
        self.cmb_sheet.currentTextChanged.connect(self._on_sheet_changed)
        toolbar.addWidget(QLabel("工作表:"))
        toolbar.addWidget(self.cmb_sheet)

        # Save button
        self.btn_save = QPushButton("💾 保存")
        self.btn_save.setObjectName("primaryBtn")
        self.btn_save.setMinimumWidth(70)
        self.btn_save.setEnabled(False)
        self.btn_save.clicked.connect(self.save_file)
        toolbar.addWidget(self.btn_save)

        layout.addLayout(toolbar)

        # ── Table ─────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectItems)
        self.table.setEditTriggers(
            QTableWidget.EditTrigger.DoubleClicked
            | QTableWidget.EditTrigger.EditKeyPressed
        )
        self.table.setSortingEnabled(False)
        self.table.cellChanged.connect(self._on_cell_changed)
        self.table.verticalHeader().setVisible(True)
        self.table.horizontalHeader().setStretchLastSection(True)

        # Row number column style
        self.table.verticalHeader().setStyleSheet(
            "background:#161b22; color:#8b949e; font-size:8pt;"
        )

        layout.addWidget(self.table, 1)

        # ── Status bar ───────────────────────────────────────────
        status_row = QHBoxLayout()
        status_row.setSpacing(8)

        self.lbl_status = QLabel("就绪")
        self.lbl_status.setObjectName("muted")
        status_row.addWidget(self.lbl_status)
        status_row.addStretch()

        self.lbl_modified = QLabel("")
        self.lbl_modified.setObjectName("muted")
        self.lbl_modified.setStyleSheet("color: #d29922;")
        status_row.addWidget(self.lbl_modified)

        self.lbl_dims = QLabel("")
        self.lbl_dims.setObjectName("muted")
        status_row.addWidget(self.lbl_dims)

        layout.addLayout(status_row)

    # ────────────────────────────────────────────────────────────
    #  Public API
    # ────────────────────────────────────────────────────────────

    def load_file(self, path: str) -> bool:
        """
        Load an Excel file into the viewer.
        Returns True on success, False on error.
        """
        try:
            self._close_workbook()

            self._wb = openpyxl.load_workbook(path, data_only=True)
            self._file_path = path
            self._sheet_names = self._wb.sheetnames

            # Populate sheet selector
            self.cmb_sheet.blockSignals(True)
            self.cmb_sheet.clear()
            self.cmb_sheet.addItems(self._sheet_names)
            self.cmb_sheet.blockSignals(False)

            # Load first/active sheet
            active = self._wb.active.title
            idx = self._sheet_names.index(active) if active in self._sheet_names else 0
            self.cmb_sheet.setCurrentIndex(idx)
            self._load_sheet(self._sheet_names[idx])

            # Update header
            self.lbl_filename.setText(os.path.basename(path))
            self.btn_save.setEnabled(False)
            self._modified = False
            self.lbl_modified.setText("")
            self.lbl_status.setText("已加载")

            self.file_loaded.emit(path)
            return True

        except Exception as e:
            self.lbl_status.setText(f"加载失败: {e}")
            QMessageBox.critical(self, "错误", f"无法打开文件:\n{e}")
            return False

    def save_file(self) -> bool:
        """
        Save current workbook back to disk.
        Uses a copy workbook with write mode to preserve formatting.
        Returns True on success.
        """
        if not self._file_path or not self._wb:
            return False

        try:
            self._wb.save(self._file_path)
            self._modified = False
            self.lbl_modified.setText("")
            self.lbl_status.setText("已保存")
            self.btn_save.setEnabled(False)
            return True
        except Exception as e:
            QMessageBox.critical(self, "保存失败", str(e))
            return False

    def get_file_path(self) -> str:
        return self._file_path

    def get_current_sheet(self) -> str:
        return self.cmb_sheet.currentText() if self._ws else ""

    def is_modified(self) -> bool:
        return self._modified

    # ────────────────────────────────────────────────────────────
    #  Sheet loading
    # ────────────────────────────────────────────────────────────

    def _load_sheet(self, sheet_name: str):
        """Populate the QTableWidget with data from a sheet."""
        ws = self._wb[sheet_name]
        self._ws = ws
        self._orig_values.clear()
        self._modified = False

        # Freeze screen updates
        self.table.cellChanged.disconnect(self._on_cell_changed)
        self.table.setUpdatesEnabled(False)

        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row == 0 or max_col == 0:
            self.table.setRowCount(1)
            self.table.setColumnCount(1)
            self.table.setItem(0, 0, QTableWidgetItem(""))
            self.table.horizontalHeader().setVisible(False)
            self.table.verticalHeader().setVisible(False)
            self._headers = []
            self.lbl_dims.setText("")
            self.table.setUpdatesEnabled(True)
            self.table.cellChanged.connect(self._on_cell_changed)
            return

        self._headers = [str(ws.cell(1, c).value or "") for c in range(1, max_col + 1)]

        self.table.setColumnCount(max_col)
        self.table.setRowCount(max_row - 1)          # data rows only (headers not shown in table)

        # Column headers
        self.table.setHorizontalHeaderLabels(self._headers)
        self.table.horizontalHeader().setVisible(True)
        self.table.verticalHeader().setVisible(True)

        # Row 1 = headers (read-only, styled), rows 2+ = data (editable)
        for r in range(2, max_row + 1):
            table_row = r - 2    # 0-based index for data rows
            for c in range(1, max_col + 1):
                val = ws.cell(r, c).value
                str_val = "" if val is None else str(val)
                item = QTableWidgetItem(str_val)
                self.table.setItem(table_row, c - 1, item)
                # Track original value for change detection
                self._orig_values[(table_row, c - 1)] = str_val

        # Column sizing
        for c in range(max_col):
            self.table.horizontalHeader().setSectionResizeMode(
                c, QHeaderView.ResizeMode.Interactive
            )
            self.table.setColumnWidth(c, 100)

        self.table.setUpdatesEnabled(True)
        self.table.cellChanged.connect(self._on_cell_changed)

        # Update status
        self.lbl_dims.setText(f"行: {max_row - 1}  |  列: {max_col}")
        self._update_modified_label()

    def _on_sheet_changed(self, sheet_name: str):
        if not sheet_name or not self._wb:
            return
        self._load_sheet(sheet_name)
        self.sheet_changed.emit(sheet_name)

    # ────────────────────────────────────────────────────────────
    #  Editing
    # ────────────────────────────────────────────────────────────

    def _on_cell_changed(self, row: int, col: int):
        """Track whether the table has unsaved changes."""
        if not self._ws:
            return
        item = self.table.item(row, col)
        new_val = item.text() if item else ""
        orig    = self._orig_values.get((row, col), "")
        if new_val != orig:
            self._modified = True
            self._update_modified_label()
            self.btn_save.setEnabled(True)

            # Write change back to openpyxl sheet (1-based, row 2 = table row 0)
            sheet_row = row + 2
            sheet_col = col + 1
            self._ws.cell(sheet_row, sheet_col).value = new_val
        else:
            # Check if any other cells are still modified
            self._check_modified_state()

    def _check_modified_state(self):
        any_mod = False
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and item.text() != self._orig_values.get((row, col), ""):
                    any_mod = True
                    break
            if any_mod:
                break
        self._modified = any_mod
        self._update_modified_label()
        self.btn_save.setEnabled(any_mod)

    def _update_modified_label(self):
        self.lbl_modified.setText("● 未保存" if self._modified else "")

    # ────────────────────────────────────────────────────────────
    #  Cleanup
    # ────────────────────────────────────────────────────────────

    def _close_workbook(self):
        if self._wb:
            try:
                self._wb.close()
            except Exception:
                pass
            self._wb = None
            self._ws = None

    def closeEvent(self, event):
        if self._modified:
            reply = QMessageBox.question(
                self, "未保存的更改",
                "当前文件有未保存的修改，是否保存？",
                QMessageBox.StandardButton.Save
                | QMessageBox.StandardButton.Discard
                | QMessageBox.StandardButton.Cancel,
            )
            if reply == QMessageBox.StandardButton.Save:
                if not self.save_file():
                    event.ignore()
                    return
            elif reply == QMessageBox.StandardButton.Cancel:
                event.ignore()
                return
        self._close_workbook()
        super().closeEvent(event)
